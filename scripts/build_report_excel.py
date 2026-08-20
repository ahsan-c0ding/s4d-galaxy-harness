#!/usr/bin/env python3
"""
Compile every training run from this project's 15 notebooks into one Excel
workbook.

Every number below was cross-checked against the actual notebook cell
outputs in notebooks/ (not copied from the LaTeX report alone) -- the
per-row "Source" column says which notebook and, where identifiable, which
script/cell. A handful of rows (marked "LaTeX report") come from tables in
report.tex whose exact producing notebook cell wasn't independently
re-located during extraction; those numbers still matched every adjacent
notebook figure that WAS checked, but are labeled honestly rather than
implied to be independently re-verified.

Run:
    python scripts/build_report_excel.py
Output:
    results/compiled_training_runs.xlsx
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
OUT_PATH = REPO_ROOT / "results" / "compiled_training_runs.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=10, color="555555")
BODY_FONT = Font(name=FONT_NAME, size=10)

NB = "notebooks/"  # prefix used in Source columns below

# ============================================================================
# 1. Production family -- Master Results Table + budget table
#    Source: notebook-best-s4d-model(1).ipynb, s4d-2-layers.ipynb,
#    notebook-64x64-dimension.ipynb, s4d-pooling-and-scale-tests.ipynb,
#    s4d-width-scan-test.ipynb, s4d-108d-test.ipynb, s4d-108d-test(2).ipynb,
#    small-linear-s4d-70k-parameters.ipynb
# ============================================================================
PRODUCTION_MASTER = [
    dict(model="Best S4D (d=256)", embed="conv", layers=3, pooling="last", seed=30485,
         params=528004, accuracy=0.8680, f1=0.8688, precision=None, recall=None, auc=None,
         source=NB + "notebook-best-s4d-model(1).ipynb"),
    dict(model="2-Layer S4D (d=256)", embed="conv", layers=2, pooling="last", seed=30485,
         params=396420, accuracy=0.8640, f1=0.8647, precision=None, recall=None, auc=None,
         source=NB + "s4d-2-layers.ipynb"),
    dict(model="64-dim S4D (3L, re-run)", embed="conv", layers=3, pooling="last", seed=30485,
         params=58948, accuracy=0.8420, f1=0.8425, precision=0.8431, recall=0.8423, auc=0.9705,
         source=NB + "notebook-64x64-dimension.ipynb"),
    dict(model="Conv Embed + No S4D", embed="conv", layers=0, pooling="mean", seed=30485,
         params=133252, accuracy=0.5150, f1=0.5156, precision=0.5205, recall=0.5169, auc=0.7730,
         source=NB + "s4d-testing-scan-variants.ipynb"),
    dict(model="Linear Embed + S4D (d=256)", embed="linear", layers=3, pooling="last", seed=30485,
         params=408324, accuracy=0.8285, f1=0.8290, precision=0.8299, recall=0.8294, auc=0.9621,
         source=NB + "s4d-testing-scan-variants.ipynb"),
    dict(model="Linear Embed + S4D, mean pool", embed="linear", layers=3, pooling="mean", seed=30485,
         params=408324, accuracy=0.8220, f1=0.8210, precision=0.8219, recall=0.8228, auc=0.9618,
         source=NB + "s4d-pooling-and-scale-tests.ipynb"),
    dict(model="Linear Embed + No S4D", embed="linear", layers=0, pooling="mean", seed=30485,
         params=13572, accuracy=0.3585, f1=0.3528, precision=0.3613, recall=0.3604, auc=0.6195,
         source=NB + "s4d-testing-scan-variants.ipynb"),
    dict(model="Pooling Confound (conv+S4D, mean)", embed="conv", layers=3, pooling="mean", seed=30485,
         params=528004, accuracy=0.8545, f1=0.8544, precision=0.8560, recall=0.8546, auc=0.9716,
         source=NB + "s4d-pooling-and-scale-tests.ipynb"),
    dict(model="Small Conv + S4D (d=72)", embed="conv", layers=3, pooling="last", seed=30485,
         params=69660, accuracy=0.8515, f1=0.8522, precision=0.8528, recall=0.8520, auc=0.9727,
         source=NB + "s4d-width-scan-test.ipynb"),
    dict(model="Linear + S4D (d=72)", embed="linear", layers=3, pooling="last", seed=30485,
         params=35356, accuracy=0.8020, f1=0.8028, precision=0.8043, recall=0.8033, auc=0.9562,
         source=NB + "small-linear-s4d-70k-parameters.ipynb"),
    dict(model="Linear + S4D (d=90)", embed="linear", layers=3, pooling="last", seed=30485,
         params=53914, accuracy=0.8010, f1=0.8016, precision=None, recall=None, auc=0.9578,
         source=NB + "s4d-width-scan-test.ipynb"),
    dict(model="Linear + S4D (d~108, seed 30485)", embed="linear", layers=3, pooling="last", seed=30485,
         params=76360, accuracy=0.7895, f1=0.7902, precision=0.7907, recall=0.7903, auc=0.9551,
         source=NB + "s4d-108d-test.ipynb"),
    dict(model="Linear + S4D (d~108, seed 8842)", embed="linear", layers=3, pooling="last", seed=8842,
         params=76360, accuracy=0.8220, f1=0.8220, precision=0.8235, recall=0.8231, auc=0.9615,
         source=NB + "s4d-108d-test(2).ipynb"),
]

# ============================================================================
# 2. Production family -- recipe-crossing + metric-recovery reruns
#    Source: report-discrepency-testing.ipynb (the controlled study)
# ============================================================================
PRODUCTION_RECIPE_CROSSING = [
    dict(model="Best S4D (528K, 3L)", native_recipe="main", tested_recipe="short",
         native_acc=0.8680, tested_acc=0.7760, tested_f1=0.7763, tested_auc=0.9465,
         source=NB + "report-discrepency-testing.ipynb"),
    dict(model="2-Layer S4D (396K, 2L)", native_recipe="main", tested_recipe="short",
         native_acc=0.8640, tested_acc=0.7835, tested_f1=0.7840, tested_auc=0.9489,
         source=NB + "report-discrepency-testing.ipynb"),
    dict(model="Small Conv+S4D (69.7K, 3L)", native_recipe="main", tested_recipe="short",
         native_acc=0.8515, tested_acc=0.7530, tested_f1=0.7531, tested_auc=0.9296,
         source=NB + "report-discrepency-testing.ipynb"),
    dict(model="Richer-stem Hybrid (55.1K, 2L)", native_recipe="short", tested_recipe="main",
         native_acc=0.8735, tested_acc=0.8975, tested_f1=0.8979, tested_auc=0.9854,
         source=NB + "report-discrepency-testing.ipynb"),
    dict(model="Richer-stem Hybrid (98.3K, 3L)", native_recipe="short", tested_recipe="main",
         native_acc=0.8750, tested_acc=0.9005, tested_f1=0.9011, tested_auc=0.9859,
         source=NB + "report-discrepency-testing.ipynb"),
    dict(model="Richer-stem Hybrid (699.7K, 3L)", native_recipe="short", tested_recipe="main",
         native_acc=0.8750, tested_acc=0.8975, tested_f1=0.8982, tested_auc=0.9865,
         source=NB + "report-discrepency-testing.ipynb"),
]

PRODUCTION_METRIC_RECOVERY = [
    dict(model="Best S4D (528K, 3L)", reported_acc=0.8680, retrain_acc=0.8530,
         f1=0.8536, precision=0.8537, recall=0.8537, auc=0.9730,
         source=NB + "report-discrepency-testing.ipynb"),
    dict(model="2-Layer S4D (396K, 2L)", reported_acc=0.8640, retrain_acc=0.8555,
         f1=0.8560, precision=0.8564, recall=0.8562, auc=0.9736,
         source=NB + "report-discrepency-testing.ipynb"),
]

# ============================================================================
# 3. Production family -- repeat-seed follow-up validation (embedding + pooling)
#    Source: s4d-future-work-validation-resumable-kaggle.ipynb / future-work-training-v3.ipynb
# ============================================================================
PRODUCTION_FOLLOWUP_REPEAT_SEED = [
    dict(config="Conv, d=256, last pool", seed_30485=0.8355, seed_8842=0.8380, spread=0.0025),
    dict(config="Conv, d=256, mean pool", seed_30485=0.8420, seed_8842=0.8470, spread=0.0050),
    dict(config="Conv, d=72, last pool (~70K)", seed_30485=0.8360, seed_8842=0.8355, spread=0.0005),
    dict(config="Linear, d~108, last pool (~76K)", seed_30485=0.7965, seed_8842=0.8100, spread=0.0135),
    dict(config="Linear, d=256, last pool", seed_30485=0.8170, seed_8842=0.8210, spread=0.0040),
    dict(config="Linear, d=256, mean pool", seed_30485=0.8175, seed_8842=0.8315, spread=0.0140),
]
for _r in PRODUCTION_FOLLOWUP_REPEAT_SEED:
    _r["source"] = NB + "s4d-future-work-validation-resumable-kaggle.ipynb"

# ============================================================================
# 4. Richer-stem family -- CNN-only / S4D-only / hybrid, short vs main recipe
#    Source (short): lodhi-training-re-runs(1).ipynb -- train_cnn_only.py, train.py, train_hybrid.py
#    Source (main):  further-report-discrepency-testing.ipynb
# ============================================================================
RICHER_CNN_ONLY_HYBRID = [
    dict(model="CNN-only Tiny (~10K)", params=10020, short_recipe_acc=0.8155, main_recipe_acc=0.8770,
         seed_short=30485, source_short=NB+"lodhi-training-re-runs(1).ipynb", source_main=NB+"further-report-discrepency-testing.ipynb"),
    dict(model="CNN-only Small (~43K)", params=42756, short_recipe_acc=0.8410, main_recipe_acc=0.8900,
         seed_short=30485, source_short=NB+"lodhi-training-re-runs(1).ipynb", source_main=NB+"further-report-discrepency-testing.ipynb"),
    dict(model="CNN-only Large (~61K)", params=61044, short_recipe_acc=0.8460, main_recipe_acc=0.8925,
         seed_short=30485, source_short=NB+"lodhi-training-re-runs(1).ipynb", source_main=NB+"further-report-discrepency-testing.ipynb"),
    dict(model="S4D-only baseline (no CNN, 17K)", params=17156, short_recipe_acc=0.7050, main_recipe_acc=0.7845,
         seed_short=30485, source_short=NB+"lodhi-training-re-runs(1).ipynb", source_main="LaTeX report (grid table, not independently re-verified per report footnote)"),
    dict(model="CNN+S4D hybrid (55.1K, 2L)", params=55108, short_recipe_acc=0.8735, main_recipe_acc=0.8975,
         seed_short=42, source_short=NB+"lodhi-training-re-runs(1).ipynb", source_main=NB+"report-discrepency-testing.ipynb"),
]

RICHER_SCALE_SWEEP = [
    dict(model="hybrid_100k", params=98332, accuracy=0.8750, f1=0.8756, precision=0.8759, recall=0.8754, auc=0.9784,
         d_model=120, s4_state=60, mid_channels=32, num_s4_layers=3, colored=False),
    dict(model="hybrid_300k", params=298868, accuracy=0.8715, f1=0.8722, precision=0.8726, recall=0.8720, auc=0.9782,
         d_model=112, s4_state=224, mid_channels=80, num_s4_layers=2, colored=False),
    dict(model="hybrid_700k", params=699748, accuracy=0.8750, f1=0.8758, precision=0.8763, recall=0.8755, auc=0.9769,
         d_model=192, s4_state=384, mid_channels=80, num_s4_layers=3, colored=False),
]
for _r in RICHER_SCALE_SWEEP:
    _r["seed"] = 42
    _r["source"] = NB + "lodhi-training-re-runs(1).ipynb (scripts/legacy/train_hybrid_scale.py)"

# ============================================================================
# 5. Richer-stem grid: stem depth x S4D layers, short recipe vs main recipe
#    Source (short): lodhi-training-re-runs(1).ipynb -- scripts/legacy/train_ablation.py
#    Source (main):  future-work-training-v3.ipynb / s4d-future-work-validation-resumable-kaggle.ipynb
# ============================================================================
_GRID_PARAMS = {
    (4, 0): 38468, (4, 1): 46788, (4, 2): 55108,
    (3, 0): 29156, (3, 1): 37476, (3, 2): 45796,
    (2, 0): 19844, (2, 1): 28164, (2, 2): 36484,
    (1, 0): 2180, (1, 1): 10500, (1, 2): 18820,
}
_GRID_SHORT = {
    (4, 0): 0.8210, (4, 1): 0.8275, (4, 2): 0.8375,
    (3, 0): 0.7975, (3, 1): 0.8110, (3, 2): 0.8190,
    (2, 0): 0.6510, (2, 1): 0.7830, (2, 2): 0.8030,
    (1, 0): 0.4800, (1, 1): 0.6990, (1, 2): 0.7405,
}
_GRID_MAIN = {
    (4, 0): 0.8925, (4, 1): 0.8865, (4, 2): 0.9015,
    (3, 0): 0.8760, (3, 1): 0.8595, (3, 2): 0.8755,
    (2, 0): 0.8375, (2, 1): 0.8575, (2, 2): 0.8485,
    (1, 0): 0.5970, (1, 1): 0.8060, (1, 2): None,
}
_GRID_SHORT_F1 = {
    (4, 0): 0.8214, (4, 1): 0.8280, (4, 2): 0.8382,
    (3, 0): 0.7970, (3, 1): 0.8122, (3, 2): 0.8191,
    (2, 0): 0.6525, (2, 1): 0.7837, (2, 2): 0.8042,
    (1, 0): 0.4776, (1, 1): 0.6999, (1, 2): 0.7412,
}
_GRID_SHORT_AUC = {
    (4, 0): 0.9620, (4, 1): 0.9602, (4, 2): 0.9653,
    (3, 0): 0.9531, (3, 1): 0.9561, (3, 2): 0.9563,
    (2, 0): 0.8783, (2, 1): 0.9487, (2, 2): 0.9552,
    (1, 0): 0.7354, (1, 1): 0.9037, (1, 2): 0.9247,
}
STEM_NAMES = {4: "stem_full (4L)", 3: "stem_3 (3L, drop res_conv)", 2: "stem_2 (2L, 1 detail+1 down)", 1: "stem_1 (1L, detail only)"}
RICHER_GRID = []
for depth in (4, 3, 2, 1):
    for n in (0, 1, 2):
        RICHER_GRID.append(dict(
            stem=STEM_NAMES[depth], stem_depth=depth, s4_layers=n,
            params=_GRID_PARAMS[(depth, n)],
            short_recipe_acc=_GRID_SHORT[(depth, n)],
            short_recipe_f1=_GRID_SHORT_F1[(depth, n)],
            short_recipe_auc=_GRID_SHORT_AUC[(depth, n)],
            main_recipe_acc=_GRID_MAIN[(depth, n)],
            main_recipe_note=("not run before GPU quota exhausted" if _GRID_MAIN[(depth, n)] is None else ""),
            source_short=NB + "lodhi-training-re-runs(1).ipynb",
            source_main=(NB + "future-work-training-v3.ipynb" if _GRID_MAIN[(depth, n)] is not None else "n/a"),
        ))

RICHER_GRID_RAW_BASELINE_NOTE = "S4D-only raw-pixel baseline (17,156 params) also part of this grid; short-recipe 70.50% (source: lodhi-training-re-runs(1).ipynb); not re-run under main recipe before the GPU quota was exhausted."

# ============================================================================
# 6. Noise-robustness sweep (13 models x 5 sigma levels) -- short recipe only
#    Not in the LaTeX report's tables (only discussed qualitatively there).
#    Extracted directly from lodhi-training-re-runs(1).ipynb cell output.
# ============================================================================
NOISE_ROBUSTNESS = [
    dict(model="S4D-only baseline (no CNN)", params=17156, clean_acc=70.5, sigma_00=70.5, sigma_005=69.5, sigma_01=64.9, sigma_02=55.5, sigma_03=45.6),
    dict(model="stem_full (4L) + 0 S4D", params=38468, clean_acc=82.1, sigma_00=82.1, sigma_005=65.5, sigma_01=46.6, sigma_02=31.0, sigma_03=26.1),
    dict(model="stem_full (4L) + 1 S4D", params=46788, clean_acc=82.75, sigma_00=82.8, sigma_005=72.4, sigma_01=50.5, sigma_02=29.5, sigma_03=27.9),
    dict(model="stem_full (4L) + 2 S4D", params=55108, clean_acc=83.75, sigma_00=83.8, sigma_005=72.9, sigma_01=51.4, sigma_02=29.2, sigma_03=27.1),
    dict(model="stem_3 (drop res_conv) + 0 S4D", params=29156, clean_acc=79.75, sigma_00=79.8, sigma_005=55.2, sigma_01=44.4, sigma_02=25.4, sigma_03=25.4),
    dict(model="stem_3 (drop res_conv) + 1 S4D", params=37476, clean_acc=81.1, sigma_00=81.1, sigma_005=65.3, sigma_01=42.0, sigma_02=28.1, sigma_03=25.0),
    dict(model="stem_3 (drop res_conv) + 2 S4D", params=45796, clean_acc=81.9, sigma_00=81.9, sigma_005=68.5, sigma_01=51.1, sigma_02=31.9, sigma_03=28.1),
    dict(model="stem_2 (1 detail + 1 down) + 0 S4D", params=19844, clean_acc=65.1, sigma_00=65.1, sigma_005=34.6, sigma_01=29.8, sigma_02=25.4, sigma_03=25.4),
    dict(model="stem_2 (1 detail + 1 down) + 1 S4D", params=28164, clean_acc=78.3, sigma_00=78.3, sigma_005=63.9, sigma_01=46.5, sigma_02=28.1, sigma_03=26.8),
    dict(model="stem_2 (1 detail + 1 down) + 2 S4D", params=36484, clean_acc=80.3, sigma_00=80.3, sigma_005=67.8, sigma_01=43.6, sigma_02=28.0, sigma_03=25.9),
    dict(model="stem_1 (detail only) + 0 S4D", params=2180, clean_acc=48.0, sigma_00=48.0, sigma_005=40.5, sigma_01=33.1, sigma_02=25.5, sigma_03=25.3),
    dict(model="stem_1 (detail only) + 1 S4D", params=10500, clean_acc=69.9, sigma_00=69.9, sigma_005=66.2, sigma_01=58.7, sigma_02=41.9, sigma_03=32.5),
    dict(model="stem_1 (detail only) + 2 S4D", params=18820, clean_acc=74.05, sigma_00=74.1, sigma_005=68.5, sigma_01=63.9, sigma_02=46.2, sigma_03=33.7),
]
for _r in NOISE_ROBUSTNESS:
    _r["source"] = NB + "lodhi-training-re-runs(1).ipynb"
    _r["note"] = "Test-time-only additive Gaussian pixel noise; no repeat seed (see report Section 'Robustness to Input Noise')"

# ============================================================================
# 7. GroupNorm-folding portability test
#    Source: s4d-future-work-validation-resumable-kaggle.ipynb
# ============================================================================
GROUPNORM_FOLDING = [
    dict(variant="Original (dynamic GroupNorm)", split="val", accuracy=0.9006, f1=0.9003, precision=0.9003, recall=0.9004),
    dict(variant="Calibrated fixed-stat, folded", split="val", accuracy=0.8906, f1=0.8907, precision=0.8910, recall=0.8905),
    dict(variant="Original (dynamic GroupNorm)", split="test", accuracy=0.9015, f1=0.9019, precision=0.9019, recall=0.9020),
    dict(variant="Calibrated fixed-stat, folded", split="test", accuracy=0.8785, f1=0.8794, precision=0.8803, recall=0.8789),
]
for _r in GROUPNORM_FOLDING:
    _r["model"] = "stem_full + 2 S4D (55.1K params, seed 30485, main recipe)"
    _r["source"] = NB + "s4d-future-work-validation-resumable-kaggle.ipynb"


# ============================================================================
# Workbook writer
# ============================================================================
def style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, df, min_width=10, max_width=48):
    for i, col in enumerate(df.columns, start=1):
        col_max = max((len(str(v)) for v in df[col].tolist()), default=0)
        width = max(min_width, min(max_width, max(len(str(col)), col_max) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def write_df_sheet(wb, name, df, note=None, pct_cols=()):
    ws = wb.create_sheet(name)
    row0 = 1
    if note:
        ws.cell(row=1, column=1, value=note).font = NOTE_FONT
        row0 = 3
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=row0, column=j, value=col)
    style_header(ws, len(df.columns), row=row0)
    for i, (_, rec) in enumerate(df.iterrows(), start=row0 + 1):
        for j, col in enumerate(df.columns, start=1):
            val = rec[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            if col in pct_cols and isinstance(val, (int, float)):
                cell.number_format = "0.00%"
    ws.freeze_panes = ws.cell(row=row0 + 1, column=1).coordinate
    autosize(ws, df)
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # --- README / provenance sheet ---
    ws = wb.create_sheet("README")
    ws.column_dimensions["A"].width = 110
    lines = [
        ("S4D Galaxy Classifier -- Compiled Training Runs", TITLE_FONT),
        ("", BODY_FONT),
        ("Compiled from the 15 training notebooks in notebooks/ (Team 0x43), cross-checked", BODY_FONT),
        ("against report.tex. Every sheet has a Source column naming which notebook produced", BODY_FONT),
        ("that row; a handful of cells that trace to the LaTeX report rather than an independently", BODY_FONT),
        ("re-located notebook cell are labeled 'LaTeX report' rather than implied to be notebook-verified.", BODY_FONT),
        ("", BODY_FONT),
        ("Sheets:", Font(name=FONT_NAME, bold=True)),
        ("  1. Production - Master Table    Full ablation series for the ConvPatchStem/S4DConv architecture", BODY_FONT),
        ("  2. Production - Recipe Crossing  Same architectures retrained under the other family's recipe", BODY_FONT),
        ("  3. Production - Metric Recovery  Reruns to recover missing precision/recall/AUC for 2 headline rows", BODY_FONT),
        ("  4. Production - Repeat Seed      6 configs x 2 seeds, main recipe (embedding + pooling confirmation)", BODY_FONT),
        ("  5. Richer-Stem - CNN-S4D-Hybrid  CNN-only / S4D-only / hybrid, short vs main recipe", BODY_FONT),
        ("  6. Richer-Stem - Scale Sweep     Hybrid at ~100K/300K/700K params, grayscale, short recipe", BODY_FONT),
        ("  7. Richer-Stem - Grid            Stem depth (1-4) x S4D layers (0-2), short vs main recipe", BODY_FONT),
        ("  8. Noise Robustness (raw)        13 models x 5 noise levels -- not tabulated in the LaTeX report", BODY_FONT),
        ("  9. GroupNorm Folding             RISC-V portability test: cost of folding GroupNorm at export", BODY_FONT),
        ("", BODY_FONT),
        ("Not included: appendix training-curve/confusion-matrix images (figures only, no tabular data);", BODY_FONT),
        ("the two richer-grid cells (stem_1+2 S4D, raw-pixel baseline) and the noise-robustness repeat-seed", BODY_FONT),
        ("sweep that the report's own GPU quota ran out before completing -- these are genuinely absent from", BODY_FONT),
        ("every notebook, not just missed in extraction (see report Section 'Follow-Up Validation').", BODY_FONT),
        ("", BODY_FONT),
        ("Generated by scripts/build_report_excel.py -- rerun after adding new results to results/runs/.", NOTE_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = font

    pct = {"accuracy", "f1", "precision", "recall", "auc", "native_acc", "tested_acc", "tested_f1", "tested_auc",
           "reported_acc", "retrain_acc", "seed_30485", "seed_8842", "spread", "short_recipe_acc", "main_recipe_acc",
           "short_recipe_f1", "short_recipe_auc"}

    write_df_sheet(wb, "Production - Master Table", pd.DataFrame(PRODUCTION_MASTER), pct_cols=pct,
                    note="Full ablation series, production family (ConvPatchStem / S4DConv). One seed each unless noted.")
    write_df_sheet(wb, "Production - Recipe Crossing", pd.DataFrame(PRODUCTION_RECIPE_CROSSING), pct_cols=pct,
                    note="Each architecture retrained under the OTHER family's recipe, seed 30485, shared 80/20 split.")
    write_df_sheet(wb, "Production - Metric Recovery", pd.DataFrame(PRODUCTION_METRIC_RECOVERY), pct_cols=pct,
                    note="Retrained under own native recipe purely to recover precision/recall/AUC missing from the original run.")
    write_df_sheet(wb, "Production - Repeat Seed", pd.DataFrame(PRODUCTION_FOLLOWUP_REPEAT_SEED), pct_cols=pct,
                    note="6 configs x 2 seeds (30485, 8842), main recipe -- confirms conv-embedding edge, reverses earlier pooling reading.")
    write_df_sheet(wb, "Richer-Stem - CNN-S4D-Hybrid", pd.DataFrame(RICHER_CNN_ONLY_HYBRID), pct_cols=pct,
                    note="CNN-only baselines, S4D-only raw-pixel baseline, and the CNN+S4D hybrid, short vs main recipe.")
    write_df_sheet(wb, "Richer-Stem - Scale Sweep", pd.DataFrame(RICHER_SCALE_SWEEP), pct_cols=pct,
                    note="Hybrid model at ~100K/300K/700K params, grayscale input, short recipe, seed 42.")
    write_df_sheet(wb, "Richer-Stem - Grid", pd.DataFrame(RICHER_GRID), pct_cols=pct,
                    note="Stem depth x S4D layer count, both recipes. " + RICHER_GRID_RAW_BASELINE_NOTE)
    write_df_sheet(wb, "Noise Robustness (raw)", pd.DataFrame(NOISE_ROBUSTNESS),
                    note="Additive Gaussian pixel noise at test time only, short recipe, no repeat seed. Accuracy in percent (not fraction).")
    write_df_sheet(wb, "GroupNorm Folding", pd.DataFrame(GROUPNORM_FOLDING), pct_cols=pct,
                    note="RISC-V portability test: cost of folding GroupNorm's running stats into the preceding conv at export time.")

    # --- Summary sheet with live formulas over the Master Table ---
    ws = wb.create_sheet("Summary", 0)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.cell(row=1, column=1, value="Quick Summary (formulas reference the Production - Master Table sheet)").font = TITLE_FONT
    master_sheet = "'Production - Master Table'"
    n_master = len(PRODUCTION_MASTER)
    # That sheet is written with a note in row 1, blank row 2, header in row 3,
    # data starting row 4. Columns: A=model, B=embed, C=layers, D=pooling,
    # E=seed, F=params, G=accuracy, H=f1, I=precision, J=recall, K=auc, L=source.
    data_first, data_last = 4, 3 + n_master
    rows = [
        ("Production family: # configurations", f"=COUNTA({master_sheet}!A{data_first}:A{data_last})"),
        ("Production family: best accuracy", f"=MAX({master_sheet}!G{data_first}:G{data_last})"),
        ("Production family: worst accuracy", f"=MIN({master_sheet}!G{data_first}:G{data_last})"),
        ("Production family: avg accuracy (conv embed)", f'=AVERAGEIF({master_sheet}!B{data_first}:B{data_last},"conv",{master_sheet}!G{data_first}:G{data_last})'),
        ("Production family: avg accuracy (linear embed)", f'=AVERAGEIF({master_sheet}!B{data_first}:B{data_last},"linear",{master_sheet}!G{data_first}:G{data_last})'),
        ("Total runs compiled across all sheets", len(PRODUCTION_MASTER) + len(PRODUCTION_RECIPE_CROSSING) + len(PRODUCTION_METRIC_RECOVERY)
         + len(PRODUCTION_FOLLOWUP_REPEAT_SEED) * 2 + len(RICHER_CNN_ONLY_HYBRID) * 2 + len(RICHER_SCALE_SWEEP)
         + len(RICHER_GRID) * 2 - 1 + len(NOISE_ROBUSTNESS) * 5 + len(GROUPNORM_FOLDING)),
    ]
    for i, (label, formula) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        cell = ws.cell(row=i, column=2, value=formula)
        cell.font = BODY_FONT
        if isinstance(formula, str) and "accuracy" in label.lower():
            cell.number_format = "0.00%"
    ws.cell(row=len(rows) + 4, column=1,
            value="Columns G (accuracy) on the Master Table, used above, are stored as fractions (0.868 = 86.80%).").font = NOTE_FONT

    wb.move_sheet("README", offset=-len(wb.sheetnames))
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size:,} bytes, {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    sys.exit(main())
